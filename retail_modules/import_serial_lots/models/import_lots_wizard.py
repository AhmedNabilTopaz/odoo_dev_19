import base64
import tempfile
from openpyxl import load_workbook
from odoo import models, fields, _
from odoo.exceptions import UserError


class ImportLotWizard(models.TransientModel):
    _name = 'import.lot.wizard'
    _description = 'Import Lots from Excel'

    move_id = fields.Many2one('stock.move', string='Stock Move', required=True)
    file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='Filename')

    def action_import(self):
        self.ensure_one()

        if not self.file:
            raise UserError(_("Please upload an Excel file."))

        picking = self.move_id.picking_id

        file_data = base64.b64decode(self.file)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file_data)
            tmp.seek(0)

            workbook = load_workbook(tmp.name)
            sheet = workbook.active

            seen_lots = set()

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

                location_name, lot_name, qty = row

                if not lot_name:
                    continue

                if lot_name in seen_lots:
                    raise UserError(
                        _('Duplicate Lot/Serial "%s" found on row %s.') % (lot_name, row_idx)
                    )

                seen_lots.add(lot_name)

                if '/' in location_name:
                    location_name = location_name.split('/', 1)[1].strip()

                location = self.env['stock.location'].search(
                    [('name', '=', location_name)],
                    limit=1
                )
                if not location:
                    raise UserError(_('Location "%s" not found.') % location_name)

                lot = self.env['stock.lot'].search(
                    [('name', '=', lot_name),
                     ('product_id', '=', self.move_id.product_id.id)],
                    limit=1
                )

                if not lot:
                    lot = self.env['stock.lot'].create({
                        'name': lot_name,
                        'product_id': self.move_id.product_id.id,
                    })

                self.env['stock.move.line'].create({
                    'picking_id': picking.id,
                    'location_id': location.id,
                    'location_dest_id': picking.location_dest_id.id,
                    'product_id': self.move_id.product_id.id,
                    'lot_id': lot.id,
                    'qty_done': float(qty or 0),
                    'move_id': self.move_id.id,
                    'product_uom_id': self.move_id.product_uom.id,
                })
